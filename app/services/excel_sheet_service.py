import io
from pathlib import Path

import openpyxl
import pandas as pd
from fastapi import HTTPException

from app.services.csv_service import (
    _clean_upload_file_name,
    stream_rows_to_normalized_file,
)
from app.services.temporary_upload_service import TemporaryUpload
from app.utils.responses import error_response

EXCEL_EXTENSIONS = {".xlsx", ".xls"}
MAX_EXCEL_SHEET_ROWS = 600_000


def _is_excel_file(file_name: str) -> bool:
    return Path(file_name).suffix.lower() in EXCEL_EXTENSIONS


def extract_sheet_names(file_content: bytes | str) -> list[str]:
    workbook = None
    try:
        workbook = openpyxl.load_workbook(
            filename=file_content if isinstance(file_content, str) else io.BytesIO(file_content),
            read_only=True,
            data_only=True,
        )
        sheet_names = list(workbook.sheetnames)
    except Exception:
        sheet_names = []
    finally:
        if workbook is not None:
            workbook.close()

    if sheet_names:
        return sheet_names
    if workbook is not None:
        if not sheet_names:
            raise error_response(status_code=400, detail="Excel file does not contain any sheets")

    try:
        excel_file = pd.ExcelFile(file_content if isinstance(file_content, str) else io.BytesIO(file_content))
        sheet_names = excel_file.sheet_names
        excel_file.close()
    except Exception as exc:
        raise error_response(status_code=400, detail="Excel file could not be read") from exc

    if not sheet_names:
        raise error_response(status_code=400, detail="Excel file does not contain any sheets")
    return sheet_names


def validate_sheet_selection(*, sheet_name: str, available_sheets: list[str]) -> None:
    if sheet_name not in available_sheets:
        raise error_response(status_code=400, detail="Selected sheet was not found in workbook")


def _max_rows_detail(sheet_name: str) -> str:
    return f"{sheet_name} exceeds the maximum supported row count of {MAX_EXCEL_SHEET_ROWS}"


def _stream_xlsx_selected_sheet(
    *,
    file_path: str,
    file_name: str,
    sheet_name: str,
    dest_path: Path,
) -> tuple[list[str], list[str], int, list[dict]]:
    """Stream one worksheet out to a normalized CSV.

    ``read_only`` keeps openpyxl yielding rows lazily, so the sheet is never held whole.
    """
    workbook = None
    try:
        workbook = openpyxl.load_workbook(
            filename=file_path,
            read_only=True,
            data_only=True,
        )
        sheet = workbook[sheet_name]
        return stream_rows_to_normalized_file(
            file_name,
            sheet.iter_rows(values_only=True),
            dest_path,
            max_data_rows=MAX_EXCEL_SHEET_ROWS,
            max_data_rows_detail=_max_rows_detail(sheet_name),
        )
    except Exception as exc:
        if isinstance(exc, HTTPException):
            raise
        raise error_response(status_code=400, detail=f"{sheet_name} could not be read") from exc
    finally:
        if workbook is not None:
            workbook.close()


def _stream_xls_selected_sheet(
    *,
    file_path: str,
    file_name: str,
    sheet_name: str,
    dest_path: Path,
) -> tuple[list[str], list[str], int, list[dict]]:
    """Stream one legacy .xls worksheet out to a normalized CSV.

    pandas has to decode the whole sheet to read it at all, but the XLS format caps a sheet at
    65,536 rows, so that read is bounded. Streaming the frame out row by row avoids the second,
    unbounded copy the rows used to be turned into.
    """
    try:
        dataframe = pd.read_excel(
            file_path,
            sheet_name=sheet_name,
            header=None,
            nrows=MAX_EXCEL_SHEET_ROWS + 1,
        )
    except Exception as exc:
        raise error_response(status_code=400, detail=f"{sheet_name} could not be read") from exc

    if len(dataframe.index) > MAX_EXCEL_SHEET_ROWS:
        raise error_response(status_code=400, detail=_max_rows_detail(sheet_name))

    if dataframe.empty and len(dataframe.columns) == 0:
        raise error_response(status_code=400, detail=f"{file_name} selected sheet is empty")

    frame = dataframe.astype(object).where(pd.notna(dataframe), None)
    return stream_rows_to_normalized_file(
        file_name,
        frame.itertuples(index=False, name=None),
        dest_path,
    )


def process_selected_sheet(
    *,
    file_path: str,
    file_name: str,
    sheet_name: str,
    dest_path: Path,
    available_sheets: list[str] | None = None,
) -> tuple[str, int, list[str], list[str], int, list[dict], str]:
    """Write one selected sheet to ``dest_path`` as a normalized CSV.

    The caller owns ``dest_path`` and is responsible for removing it.
    """
    clean_file_name = _clean_upload_file_name(file_name)
    if not _is_excel_file(clean_file_name):
        raise error_response(status_code=400, detail="Only XLSX and XLS files support sheet selection")
    if available_sheets is None:
        available_sheets = extract_sheet_names(file_path)
    validate_sheet_selection(sheet_name=sheet_name, available_sheets=available_sheets)

    display_file_name = f"{clean_file_name} ({sheet_name})"
    stream_sheet = (
        _stream_xlsx_selected_sheet
        if Path(clean_file_name).suffix.lower() == ".xlsx"
        else _stream_xls_selected_sheet
    )
    columns, internal_columns, total_rows, sample_rows = stream_sheet(
        file_path=file_path,
        file_name=display_file_name,
        sheet_name=sheet_name,
        dest_path=dest_path,
    )

    if not internal_columns:
        raise error_response(status_code=400, detail=f"{display_file_name} selected sheet is empty")
    if not total_rows:
        raise error_response(
            status_code=400,
            detail=f"{display_file_name} selected sheet does not contain data rows",
        )

    file_size = Path(file_path).stat().st_size
    return (
        clean_file_name,
        file_size,
        columns,
        internal_columns,
        total_rows,
        sample_rows,
        sheet_name,
    )


def process_temporary_upload_selection(
    *,
    upload: TemporaryUpload,
    sheet_name: str,
    dest_path: Path,
) -> tuple[str, int, list[str], list[str], int, list[dict], str]:
    return process_selected_sheet(
        file_path=upload.temp_file_path,
        file_name=upload.original_file_name,
        sheet_name=sheet_name,
        dest_path=dest_path,
        available_sheets=upload.available_sheets,
    )
